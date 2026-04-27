import openpyxl
from openpyxl import load_workbook
from urllib.parse import urlparse

# === CHANGE THESE ===
INPUT_FILE = 'nonprofits_with_urls.xlsx'  # Your file with the URLs
OUTPUT_FILE = 'nonprofits_cleaned_domains.xlsx'  # Output file
STATES_TO_PROCESS = ['FL']

def extract_main_domain(url):
    """Extract just the main domain from a URL"""
    if not url or url == 'Not found' or url == 'N/A':
        return url
    
    try:
        # Parse the URL
        parsed = urlparse(url)
        domain = parsed.netloc
        
        # Remove 'www.' if present
        if domain.startswith('www.'):
            domain = domain[4:]
        
        return domain
    except:
        return url  # Return original if parsing fails

# Load the workbook
print(f"Loading file: {INPUT_FILE}")
workbook = load_workbook(INPUT_FILE)

print(f"Found {len(workbook.sheetnames)} sheets\n")

# Process each sheet
for sheet_name in workbook.sheetnames:
    # Skip sheets not in our list (if you want to process all sheets, remove this check)
    if sheet_name not in STATES_TO_PROCESS:
        print(f"Skipping sheet: {sheet_name}")
        continue
    
    sheet = workbook[sheet_name]
    print(f"Processing sheet: {sheet_name}")
    
    # Find the URL column (assuming it's column E)
    url_column = 'E'
    
    # Process each row (skip header)
    cleaned_count = 0
    for row_num in range(2, sheet.max_row + 1):
        original_url = sheet[f'{url_column}{row_num}'].value
        
        if original_url and original_url != 'Not found':
            cleaned_domain = extract_main_domain(original_url)
            
            # Only update if it actually changed
            if cleaned_domain != original_url:
                sheet[f'{url_column}{row_num}'] = cleaned_domain
                cleaned_count += 1
    
    print(f"  Cleaned {cleaned_count} URLs in {sheet_name}\n")

# Save the workbook
print(f"Saving to: {OUTPUT_FILE}")
workbook.save(OUTPUT_FILE)

print(f"\n✅ Done! Cleaned domains saved to: {OUTPUT_FILE}")
print("\nExamples of what was cleaned:")
print("Before: https://www.duvalschools.org/o/bsms/page/communities-in-schools/")
print("After:  duvalschools.org")